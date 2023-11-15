import csv

import pymysql.cursors
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

import NameDB
import openpyxl
from openpyxl.utils import get_column_letter


def check_db() -> None:
    try:
        conn = pymysql.connect(host='localhost',
                               user='root',
                               password='root',
                               cursorclass=pymysql.cursors.Cursor)
        cursor = conn.cursor()
        cursor.execute("CREATE DATABASE IF NOT EXISTS `%s`" % NameDB.namedb.get_name_db())
    except pymysql.err.ProgrammingError as e:
        print(e)

    conn = pymysql.connect(host='localhost',
                           user='root',
                           password='root',
                           database=NameDB.namedb.get_name_db(),
                           cursorclass=pymysql.cursors.Cursor)
    cursor = conn.cursor()
    print("База данных подключена")

    try:
        cursor.execute("SELECT * FROM %s" % NameDB.namedb.get_name_tb())
    except pymysql.err.MySQLError:
        with open('create_structure.sql', 'r') as sql_file:
            sql_script = sql_file.read()
            cursor.execute(sql_script % NameDB.namedb.get_name_tb())
            conn.commit()
            print("Скрипт SQL успешно выполнен")
    return


def save_result(operat, res):
    try:
        conn = pymysql.connect(host='localhost',
                               user='root',
                               password='root',
                               database=NameDB.namedb.get_name_db(),
                               cursorclass=pymysql.cursors.Cursor)
        cursor = conn.cursor()
        cursor.execute("INSERT INTO " + NameDB.namedb.get_name_tb() + f" (operat, res) VALUES (%s, %s)",
                       (str(operat), str(res)))
        conn.commit()

        conn = pymysql.connect(host='localhost',
                               user='root',
                               password='root',
                               database=NameDB.namedb.get_name_db(),
                               cursorclass=pymysql.cursors.Cursor)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM " + NameDB.namedb.get_name_tb())
        print(cursor.fetchall()[-1])
    except pymysql.err.DataError as e:
        print('Ошибка с данными:', e)
    except pymysql.err.DatabaseError as e:
        print(e)
    return


def save_db_to_xlsx():
    try:
        conn = pymysql.connect(host='localhost',
                               user='root',
                               password='root',
                               database=NameDB.namedb.get_name_db(),
                               cursorclass=pymysql.cursors.Cursor)
        new_df = pd.read_sql("SELECT * FROM " + NameDB.namedb.get_name_tb(), conn)

        # Создание нового файла Excel
        wb = openpyxl.Workbook()
        ws = wb.active

        # Запись данных в таблицу Excel
        for r in dataframe_to_rows(new_df, index=False, header=True):
            ws.append(r)

        # Настройка ширины столбцов на основе длины данных
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохранение файла Excel
        file1 = input("Введите имя файла с расширением xlsx: ")
        wb.save(file1)

        print(new_df)
    except pymysql.err.DatabaseError as e:
        print(e)
    return



